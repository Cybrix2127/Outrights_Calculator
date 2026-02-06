from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from flask import after_this_request
from datetime import datetime, timedelta
import csv
import io
import os
import json
from threading import Lock

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ROOT = os.path.dirname(os.path.abspath(__file__))
CASES_FILE = os.path.join(ROOT, 'cases.json')

app = Flask(__name__, static_folder=ROOT, static_url_path='')
lock = Lock()


@app.after_request
def add_cors_headers(response):
    """Add CORS headers so the frontend served from file:// or different port works."""
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response


@app.route('/')
def index():
    return send_from_directory(ROOT, 'index.html')


def parse_rate(text):
    """Parse rate input: accepts formats like '25bps', '0.25%', '5.25', or raw number.

    Returns: float percent (e.g., 0.25 for 25bps, or 5.25 for 5.25%).
    """
    if text is None:
        return 0.0
    text = str(text).strip()
    if text == '':
        return 0.0
    # bps explicit (e.g., '25bps' or '+25bps')
    if text.lower().endswith('bps'):
        try:
            return float(text[:-3]) / 100.0
        except ValueError:
            raise ValueError(f'Invalid bps value: {text}')
    # percent explicit (e.g., '0.25%')
    if text.endswith('%'):
        try:
            return float(text[:-1])
        except ValueError:
            raise ValueError(f'Invalid percent value: {text}')
    # Plain number: treat as bps (this is the most common use case for meetings/adjustments)
    try:
        v = float(text)
    except ValueError:
        raise ValueError(f'Invalid numeric value: {text}')
    return v / 100.0


def get_last_working_day(year, month):
    """Return the last working day (Mon-Fri) datetime for the given year/month."""
    # move to last day of month
    if month == 12:
        last = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last = datetime(year, month + 1, 1) - timedelta(days=1)
    # if weekend adjust
    if last.weekday() == 5:  # Saturday
        last = last - timedelta(days=1)
    elif last.weekday() == 6:  # Sunday
        last = last - timedelta(days=2)
    return last


def compute_outrights(effr_pct, meetings, me_bps=0.0, qe_bps=0.0, ye_bps=0.0):
    """
    Compute monthly outrights for 2026 using daily averages.

    - `effr_pct` is percent (e.g., 5.25)
    - `meetings` is a list of {'date': 'YYYY-MM-DD', 'change_pct': percent}
    - `me_bps`, `qe_bps`, `ye_bps` are basis points (numeric), applied as percent = bps/100

    Meeting changes are effective from the next calendar day after the meeting.
    ME/QE/YE are applied on the last working day and carried forward through consecutive weekends (including into the next month) until the next weekday clears them.
    """
    meetings_sorted = sorted(meetings, key=lambda x: datetime.strptime(x['date'], '%Y-%m-%d'))

    month_sums = [0.0] * 12
    month_counts = [0] * 12

    start = datetime(2026, 1, 1)
    end = datetime(2026, 12, 31)

    cum_change = 0.0
    mi = 0
    special_active = False
    special_value = 0.0

    # convert bps to percent
    me_pct = float(me_bps) / 100.0
    qe_pct = float(qe_bps) / 100.0
    ye_pct = float(ye_bps) / 100.0

    d = start
    while d <= end:
        # apply meeting changes effective from next day
        while mi < len(meetings_sorted) and datetime.strptime(meetings_sorted[mi]['date'], '%Y-%m-%d') < d:
            cum_change += meetings_sorted[mi]['change_pct']
            mi += 1

        day_rate = effr_pct + cum_change

        month = d.month
        last_wd = get_last_working_day(2026, month)

        # If this is the last working day: set special_active and special_value
        if d.date() == last_wd.date():
            special_active = True
            if month == 12:
                special_value = ye_pct
            elif month in (3, 6, 9):
                special_value = qe_pct
            else:
                special_value = me_pct

        # If weekday and not last working day, clear special
        if d.weekday() < 5 and d.date() != last_wd.date():
            special_active = False
            special_value = 0.0

        if special_active:
            day_rate += special_value

        month_sums[d.month - 1] += day_rate
        month_counts[d.month - 1] += 1

        d = d + timedelta(days=1)

    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    out = []
    for i in range(12):
        avg = month_sums[i] / month_counts[i] if month_counts[i] > 0 else effr_pct
        outright = 100.0 - avg
        out.append({'month': f"{month_names[i]} 2026", 'avg_rate': round(avg, 4), 'outright': round(outright, 4)})
    return out


@app.route('/api/compute', methods=['POST', 'OPTIONS'])
def api_compute():
    if request.method == 'OPTIONS':
        return jsonify({}), 200

    data = request.get_json(force=True)
    try:
        # EFFR: parse as rate string (e.g. "5.25%" -> 5.25, "525bps" -> 5.25)
        effr_text = data.get('effr', '5.25%')
        effr_pct = parse_rate(effr_text)

        # Meetings: parse each value as rate change string
        meetings_input = data.get('meetings', {})
        meetings = []
        if isinstance(meetings_input, dict):
            for date, raw in meetings_input.items():
                change_pct = 0.0
                if raw is not None and raw != '':
                    change_pct = parse_rate(str(raw))
                meetings.append({'date': date, 'change_pct': change_pct})
        else:
            for item in meetings_input:
                meetings.append({'date': item.get('date'), 'change_pct': float(item.get('change_pct', 0))})

        # ME/QE/YE: parse as rate strings (plain numbers treated as bps)
        me_text = data.get('me', '0')
        qe_text = data.get('qe', '0')
        ye_text = data.get('ye', '0')
        # parse_rate returns percent, but compute_outrights expects bps input
        # so we need to convert: parse_rate("25") -> 0.25%, then *100 = 25 bps
        me_bps = parse_rate(str(me_text)) * 100.0
        qe_bps = parse_rate(str(qe_text)) * 100.0
        ye_bps = parse_rate(str(ye_text)) * 100.0

        results = compute_outrights(effr_pct, meetings, me_bps, qe_bps, ye_bps)
        return jsonify({'success': True, 'data': results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/download-csv', methods=['POST', 'OPTIONS'])
def api_download_csv():
    data = request.get_json(force=True)
    try:
        resp = api_compute()
        if isinstance(resp, tuple) and resp[1] != 200:
            return resp
        body = resp.get_json()
        if not body.get('success'):
            return jsonify({'success': False, 'error': 'Compute failed'}), 400
        results = body['data']

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Month', 'Avg Rate (%)', 'Outright'])
        for r in results:
            writer.writerow([r['month'], r['avg_rate'], r['outright']])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name='outrights_2026.csv')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/download-excel', methods=['POST', 'OPTIONS'])
def api_download_excel():
    if not HAS_OPENPYXL:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 400
    
    try:
        # Load all saved cases
        cases = load_cases()
        
        if not cases:
            # If no cases saved, compute current and return single sheet
            data = request.get_json(force=True)
            resp = api_compute()
            if isinstance(resp, tuple) and resp[1] != 200:
                return resp
            body = resp.get_json()
            if not body.get('success'):
                return jsonify({'success': False, 'error': 'Compute failed'}), 400
            results = body['data']
            cases = [{'name': 'Current', 'results': results}]
        
        wb = Workbook()
        # Remove default sheet
        if len(wb.sheetnames) > 0:
            wb.remove(wb.active)
        
        # Create a sheet for each case
        for i, case in enumerate(cases):
            case_name = case.get('name', f'Case {i}')
            # Sanitize sheet name (max 31 chars, no special chars)
            safe_name = case_name.replace('[', '').replace(']', '').replace(':', '').replace('*', '').replace('?', '').replace('/', '').replace('\\', '')
            safe_name = safe_name[:31] if len(safe_name) > 31 else safe_name
            if not safe_name:
                safe_name = f'Case {i}'
            
            ws = wb.create_sheet(title=safe_name)
            results = case.get('results', [])
            
            # Headers
            ws.append(['Month', 'Avg Rate (%)', 'Outright', '1M Spread'])
            
            # Data rows with 1M spread calculation
            for j, r in enumerate(results):
                spread = 'N/A'
                if j < len(results) - 1:
                    spread = round(r['outright'] - results[j+1]['outright'], 4)
                ws.append([r['month'], r['avg_rate'], r['outright'], spread])
            
            # Auto-width columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='outrights_2026.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


def load_cases():
    if not os.path.exists(CASES_FILE):
        return []
    with lock:
        with open(CASES_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except Exception:
                return []


def save_cases(cases):
    with lock:
        with open(CASES_FILE, 'w', encoding='utf-8') as f:
            json.dump(cases, f, indent=2)


@app.route('/api/save-case', methods=['POST', 'OPTIONS'])
def api_save_case():
    data = request.get_json(force=True)
    name = data.get('name') or f"Case {int(datetime.utcnow().timestamp())}"
    resp = api_compute()
    if isinstance(resp, tuple) and resp[1] != 200:
        return resp
    body = resp.get_json()
    if not body.get('success'):
        return jsonify({'success': False, 'error': 'Compute failed'}), 400
    results = body['data']

    cases = load_cases()
    case_id = len(cases)
    case = {'id': case_id, 'name': name, 'inputs': data, 'results': results, 'created': datetime.utcnow().isoformat()}
    cases.append(case)
    save_cases(cases)
    return jsonify({'success': True, 'case': case})


@app.route('/api/update-case/<int:case_id>', methods=['POST', 'OPTIONS'])
def api_update_case(case_id):
    data = request.get_json(force=True)
    resp = api_compute()
    if isinstance(resp, tuple) and resp[1] != 200:
        return resp
    body = resp.get_json()
    if not body.get('success'):
        return jsonify({'success': False, 'error': 'Compute failed'}), 400
    results = body['data']

    cases = load_cases()
    for c in cases:
        if c['id'] == case_id:
            c['inputs'] = data
            c['results'] = results
            c['updated'] = datetime.utcnow().isoformat()
            save_cases(cases)
            return jsonify({'success': True, 'results': results})
    return jsonify({'success': False, 'error': 'Case not found'}), 404


@app.route('/api/list-cases', methods=['GET'])
def api_list_cases():
    cases = load_cases()
    meta = [{'id': c['id'], 'name': c.get('name'), 'created': c.get('created')} for c in cases]
    return jsonify({'success': True, 'cases': meta})


@app.route('/api/get-case/<int:case_id>', methods=['GET'])
def api_get_case(case_id):
    cases = load_cases()
    for c in cases:
        if c['id'] == case_id:
            return jsonify({'success': True, 'case': c})
    return jsonify({'success': False, 'error': 'Not found'}), 404


@app.route('/api/delete-case/<int:case_id>', methods=['POST', 'OPTIONS'])
def api_delete_case(case_id):
    cases = load_cases()
    found = False
    new_cases = []
    for c in cases:
        if c['id'] == case_id:
            found = True
            continue
        new_cases.append(c)
    if not found:
        return jsonify({'success': False, 'error': 'Case not found'}), 404
    # reassign ids to keep them sequential
    for idx, c in enumerate(new_cases):
        c['id'] = idx
    save_cases(new_cases)
    return jsonify({'success': True})


@app.route('/api/clear-cases', methods=['POST', 'OPTIONS'])
def api_clear_cases():
    save_cases([])
    return jsonify({'success': True})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True)
